"""
Generates the ASTAT Excel report from graded results.

Workbook layout:
  - One sheet per model  (columns: Subject, Question, Ground Truth, Response, Classification, Reasoning, Coverage, Conciseness)
  - "Summary" sheet      (pass-rate table: models vs subjects + overall)
  - "Raw Data" sheet     (flat dump of every row for further analysis)
"""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent.parent / "output" / "results.xlsx"

GRADE_COLORS = {
    "Pass": "C6EFCE",
    "False Positive": "FFEB9C",
    "Silent Failure": "FFCC99",
    "Verifiable Hallucination": "FFC7CE",
    "Refusal": "D9D9D9",
    "Ungraded": "F2F2F2",
    "Error": "E0E0FF",
}

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")

MODEL_COLS = [
    "Subject", "Question", "Ground Truth",
    "Response", "Classification", "Reasoning",
    "Key Point Coverage", "Conciseness",
]


def _set_header(ws, cols: list[str]) -> None:
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _col_widths(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_model_sheet(wb: openpyxl.Workbook, model_name: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title=model_name[:31])  # Excel sheet name limit
    _set_header(ws, MODEL_COLS)

    for r, row in enumerate(rows, 2):
        grade = row.get("classification", "Ungraded")
        fill_color = GRADE_COLORS.get(grade, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_color)

        values = [
            row.get("subject", ""),
            row.get("question", ""),
            row.get("ground_truth", ""),
            row.get("response", ""),
            grade,
            row.get("reasoning", ""),
            row.get("key_point_coverage", ""),
            row.get("conciseness", ""),
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _col_widths(ws, {1: 20, 2: 50, 3: 40, 4: 50, 5: 22, 6: 45, 7: 18, 8: 14})
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _write_summary_sheet(wb: openpyxl.Workbook, results: dict[str, list[dict]]) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    all_subjects = sorted({row["subject"] for rows in results.values() for row in rows})
    model_names = list(results.keys())

    # Header row: blank, then model names
    ws.cell(row=1, column=1, value="Subject").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    for c, name in enumerate(model_names, 2):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    def pass_rate(rows: list[dict], subject: str | None = None) -> str:
        filtered = [r for r in rows if subject is None or r["subject"] == subject]
        if not filtered:
            return "N/A"
        passed = sum(1 for r in filtered if r.get("classification") == "Pass")
        return f"{passed}/{len(filtered)} ({100*passed//len(filtered)}%)"

    for r, subject in enumerate(all_subjects, 2):
        ws.cell(row=r, column=1, value=subject).font = Font(bold=True)
        for c, model in enumerate(model_names, 2):
            ws.cell(row=r, column=c, value=pass_rate(results[model], subject))

    # Overall row
    overall_row = len(all_subjects) + 2
    ws.cell(row=overall_row, column=1, value="OVERALL").font = Font(bold=True)
    for c, model in enumerate(model_names, 2):
        ws.cell(row=overall_row, column=c, value=pass_rate(results[model]))

    _col_widths(ws, {1: 22, **{c: 24 for c in range(2, len(model_names) + 2)}})
    ws.freeze_panes = "B2"


def _write_raw_sheet(wb: openpyxl.Workbook, results: dict[str, list[dict]]) -> None:
    ws = wb.create_sheet(title="Raw Data")
    headers = ["Model"] + MODEL_COLS
    _set_header(ws, headers)

    r = 2
    for model, rows in results.items():
        for row in rows:
            values = [model] + [
                row.get("subject", ""), row.get("question", ""),
                row.get("ground_truth", ""), row.get("response", ""),
                row.get("classification", ""), row.get("reasoning", ""),
                row.get("key_point_coverage", ""), row.get("conciseness", ""),
            ]
            for c, val in enumerate(values, 1):
                ws.cell(row=r, column=c, value=val).alignment = Alignment(wrap_text=True)
            r += 1

    _col_widths(ws, {1: 22, 2: 20, 3: 45, 4: 35, 5: 45, 6: 20, 7: 40, 8: 16, 9: 14})
    ws.freeze_panes = "A2"


def generate_csv_report(results: dict[str, list[dict]], out_dir: Path = OUTPUT_PATH.parent) -> list[Path]:
    """
    Export results as CSV files importable into Google Sheets.
    Produces:
      - summary.csv        — pass-rate table (models × subjects)
      - results_all.csv    — every row from every model, flat
    Returns list of paths written.
    """
    import csv

    out_dir.mkdir(parents=True, exist_ok=True)
    written = []

    # --- summary.csv ---
    all_subjects = sorted({row["subject"] for rows in results.values() for row in rows})
    model_names = list(results.keys())

    def pass_rate(rows, subject=None):
        filtered = [r for r in rows if subject is None or r["subject"] == subject]
        if not filtered:
            return "N/A"
        passed = sum(1 for r in filtered if r.get("classification") == "Pass")
        return f"{passed}/{len(filtered)} ({100*passed//len(filtered)}%)"

    summary_path = out_dir / "summary.csv"
    with open(summary_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Subject"] + model_names)
        for subject in all_subjects:
            w.writerow([subject] + [pass_rate(results[m], subject) for m in model_names])
        w.writerow(["OVERALL"] + [pass_rate(results[m]) for m in model_names])
    written.append(summary_path)

    # --- results_all.csv ---
    all_path = out_dir / "results_all.csv"
    headers = ["Model", "Subject", "Question", "Ground Truth",
               "Response", "Classification", "Reasoning",
               "Key Point Coverage", "Conciseness"]
    with open(all_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for model, rows in results.items():
            for row in rows:
                w.writerow([
                    model,
                    row.get("subject", ""),
                    row.get("question", ""),
                    row.get("ground_truth", ""),
                    row.get("response", ""),
                    row.get("classification", ""),
                    row.get("reasoning", ""),
                    row.get("key_point_coverage", ""),
                    row.get("conciseness", ""),
                ])
    written.append(all_path)
    return written


def generate_report(results: dict[str, list[dict]], path: Path = OUTPUT_PATH) -> Path:
    """
    results: { model_name: [ {subject, question, ground_truth, response, classification, ...} ] }
    Writes Excel workbook and returns the output path.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    _write_summary_sheet(wb, results)
    for model_name, rows in results.items():
        _write_model_sheet(wb, model_name, rows)
    _write_raw_sheet(wb, results)

    wb.save(path)
    return path
